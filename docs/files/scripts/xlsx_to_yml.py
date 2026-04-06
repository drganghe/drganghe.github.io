#!/usr/bin/env python3
"""
xlsx_to_yml.py  —  Convert publications.xlsx to publications.yml
                   for use as a Quarto custom listing data source.

Usage:
    python xlsx_to_yml.py                          # uses default paths below
    python xlsx_to_yml.py input.xlsx output.yml    # custom paths

Requirements:
    pip install openpyxl
"""

import sys
from pathlib import Path
import openpyxl

# ── Default file paths (edit here or pass as command-line args) ──────────────
DEFAULT_INPUT  = "publications.xlsx"
DEFAULT_OUTPUT = "publications.yml"

# ── Column header → YAML key mapping ────────────────────────────────────────
KEY_MAP = {
    "Section":               "section",
    "Authors":               "authors",
    "Year":                  "year",
    "Date":                  "date",
    "Title":                 "title",
    "Paper Link":            "path",    
    "Journal":               "journal",
    "Volume":                "volume",
    "Issue":                 "issue",
    "Pages":                 "pages",
    "DOI":                   "doi",
    "PDF":                   "pdf",
    "Preprint":              "preprint",
    "ShareIt":               "shareit",
    "Supplemental Information": "supplemental",
    "GitHub":                 "github",
    "Code":                 "code",
    "Data":                 "data",
    "Highly Cited":         "highlycited",
    "Hot Paper":              "hotpaper",
    "Awards":                "awards",
    "Media Coverage":        "mediacoverage",
    "Invited Presentation":  "invitedpresentation",
    "Categories":            "categories",
}

# Keys whose values should be cast to int when possible
INT_KEYS = {"year"}


def yaml_scalar(value: str) -> str:
    """Return a properly quoted YAML scalar string for a given value."""
    # Characters that require quoting in YAML
    SPECIAL = set(':#|>[]{}*!,%@`')
    if any(c in value for c in SPECIAL) or value.startswith(('"', "'", '-', '?')):
        escaped = value.replace('\\', '\\\\').replace('"', '\\"')
        return f'"{escaped}"'
    return value


def yaml_inline_list(values: list[str]) -> str:
    """Return a YAML inline list: [a, b, "c d"]."""
    return "[" + ", ".join(yaml_scalar(v) for v in values) + "]"


def parse_categories(value: str) -> list[str]:
    """Parse a categories cell into a list for Quarto listing."""
    s = str(value).strip()
    if not s:
        return []

    # Common separators in spreadsheets
    for sep in (";", "|", ","):
        s = s.replace(sep, ",")

    items = [x.strip() for x in s.split(",")]
    return [x for x in items if x]


def parse_row(headers: list, row: tuple) -> dict:
    """Convert a spreadsheet row into an ordered dict of YAML fields."""
    rec = {}
    for header, cell_value in zip(headers, row):
        key = KEY_MAP.get(header, header.lower().replace(" ", "-"))
        if cell_value is None:
            continue
        if key in INT_KEYS:
            try:
                rec[key] = int(cell_value)
            except (ValueError, TypeError):
                s = str(cell_value).strip()
                if s:
                    rec[key] = s
        elif key == "categories":
            cats = parse_categories(cell_value)
            if cats:
                rec[key] = cats
        else:
            s = str(cell_value).strip()
            if s:
                rec[key] = s
    return rec


def record_to_yaml(rec: dict) -> str:
    """Serialise one publication record to a YAML list item block."""
    # Ordered field output
    FIELD_ORDER = [
        "section", "authors", "year", "date", "title", "path", "journal",
        "volume", "issue", "pages", "doi", "pdf", "preprint", "shareit", "supplemental",
        "github", "code", "data", "highlycited", "hotpaper", "awards", "mediacoverage", "invitedpresentation",
        "categories",
    ]
    lines = []
    first = True
    for key in FIELD_ORDER:
        if key not in rec:
            continue
        value = rec[key]
        if first:
            prefix = "- "
            indent = "  "
            first = False
        else:
            prefix = "  "
            indent = "  "

        if isinstance(value, int):
            lines.append(f"{prefix}{key}: {value}")
        elif isinstance(value, list) and key == "categories":
            lines.append(f"{prefix}{key}: {yaml_inline_list([str(v).strip() for v in value if str(v).strip()])}")
        else:
            lines.append(f"{prefix}{key}: {yaml_scalar(value)}")

    return "\n".join(lines)


def convert(input_path: str, output_path: str) -> None:
    print(f"Reading  : {input_path}")
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    print(f"Columns  : {headers}")

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip completely empty rows
        if all(v is None for v in row):
            continue
        rec = parse_row(headers, row)
        if rec:
            records.append(rec)

    print(f"Records  : {len(records)}")

    yaml_blocks = [record_to_yaml(r) for r in records]
    output_text = "\n\n".join(yaml_blocks) + "\n"

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(output_text)

    print(f"Written  : {output_path}")

    # Quick validation: re-parse with PyYAML if available
    try:
        import yaml
        with open(output_path, encoding="utf-8") as f:
            parsed = yaml.safe_load(f)
        print(f"Validated: {len(parsed)} entries parsed successfully by PyYAML")
    except ImportError:
        print("Tip: install PyYAML (`pip install pyyaml`) for automatic validation")
    except Exception as e:
        print(f"WARNING: YAML validation failed — {e}")


def should_convert(input_path: str, output_path: str, force: bool = False) -> bool:
    """Return True only when input is newer than output (or output missing)."""
    if force:
        return True

    src = Path(input_path)
    dst = Path(output_path)

    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if not dst.exists():
        print(f"Output missing: {output_path}; running conversion.")
        return True

    src_mtime = src.stat().st_mtime
    dst_mtime = dst.stat().st_mtime
    if src_mtime > dst_mtime:
        print(f"Detected update in {input_path}; running conversion.")
        return True

    print(f"No update in {input_path}; skip conversion.")
    return False


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if a != "--force"]
    force = "--force" in sys.argv[1:]

    input_file = args[0] if len(args) > 0 else DEFAULT_INPUT
    output_file = args[1] if len(args) > 1 else DEFAULT_OUTPUT

    if should_convert(input_file, output_file, force=force):
        convert(input_file, output_file)